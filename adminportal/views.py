from csv import excel
from pkg_resources import working_set
from rest_framework import generics
from rest_framework.parsers import MultiPartParser
from django.contrib.auth.models import Group
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from database.models import *
from .serializers import *
from Allauth.serializers import RegisterSerializer
from rest_framework import status
from .permissions import *
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework import filters
from rest_framework.pagination import LimitOffsetPagination
from Allauth.serializers import *
from django.db.models import Count, Q, Case, When, IntegerField, F, Value
from django.db.models.functions import Cast, Substr, Coalesce
from django.db.models.expressions import Func
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
# Create your views here.
from .permissions import IsSupervisor
# from excel_response import ExcelResponse
from datetime import datetime, timedelta
from rest_framework.decorators import api_view
from doctorsApp.permissions import IsMO
from adminportal.utils import get_aggregated_data
from knox.models import AuthToken

class CustomPageNumberPagination(PageNumberPagination):
    page_size = 10  # You can adjust this value based on your requirements
    page_size_query_param = 'page_size'
    max_page_size = 100

class PostUserGroupResquest(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = UpdateSerializer
    def post(self, request , *args, **kwargs):
        serializer = self.get_serializer(data = request.data )
        if serializer.is_valid():
            serializer.save()
            return Response({
                'status': 'success',
                'message': 'request raise successfully'
            })
        else:
            return Response(serializer.errors)

class GetGroupList(generics.GenericAPIView):
    permission_classes = [IsAuthenticated , IsSupervisor ]
    serializer_class = GroupListSerializer
    def get(self, request):
        group_list = Group.objects.all()
        serializer = self.get_serializer(group_list , many = True).data
        return Response({'group_list': serializer})

class GetGroupRequestList(generics.ListAPIView):
    serializer_class = GetGroupRequestListSerializer
    permission_classes = [IsAuthenticated , IsAdmin | IsMOH| IsViewAdmin]
    pagination_class = LimitOffsetPagination
    model = serializer_class.Meta.model
    filter_backends = (filters.SearchFilter,)


    def get_queryset(self):
        queryset = self.model.objects.filter( status=False   )

        search_terms = self.request.query_params.get('search', None )
        if search_terms:
            queryset = queryset.filter(
                Q(user__name__icontains=search_terms) |
                Q(username__icontains=search_terms) |
                Q(phoneNumber__icontains=search_terms) )

        return queryset

    def get(self, request, *args, **kwargs):

        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer.data})

        serializer = self.get_serializer(queryset, many=True)
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                        'data': serializer.data})

class updateUserGroupRequest(generics.GenericAPIView):
    serializer_class = UpdateGroupRequest
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated , IsAdmin]


    def patch(self, request ,  id ,  *args, **kwargs):
        try:
            instance = UserApprovalRecords.objects.get(id=id)
        except:
            return Response({'status': 'error',
                            'message': 'deatils not found'}, status=400)
        serializer = self.get_serializer(instance , data = request.data , partial = True )
        if serializer.is_valid():
            serializer.save()
            return Response({"status" : "success" ,
                            "message" : "User group updated successfully"
                    },status=status.HTTP_201_CREATED)
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = value[0]
            return Response({'message': error_message,
                            'status' : 'error'}, status=400)

class UserCountsAPI(APIView):
    # permission_classes = [ IsAuthenticated, IsAdmin]
    def get(self, request, *args, **kwargs):
        all = CustomUser.objects.all()
        CHV_ASHA_count = all.filter(groups__name='CHV-ASHA').count()
        MO_count = all.filter(groups__name='mo').count()
        ANM_count = all.filter(groups__name='healthworker').count()
        return Response({
            'CHV_ASHA_count' : CHV_ASHA_count ,
            'MO_count' : MO_count ,
            'ANM_count' : ANM_count
        } , status=status.HTTP_200_OK)

class InsertUsersByadmin(generics.GenericAPIView):
    # permission_classes = [permissions.IsAuthenticated,]
    serializer_class = AddUserSerializer
    # parser_classes = [MultiPartParser]
    permission_classes = (IsAuthenticated , IsAdmin | IsSupervisor | IsMOH)


    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        # print(serializer)
        try:
            if serializer.is_valid():

                group = Group.objects.get(name=serializer.validated_data.get("group"))

                user = serializer.save(is_active = True)
                customuser = serializer.validated_data
                data = RegisterSerializer(customuser, context=self.get_serializer_context()).data
                user.groups.add(group)
                addSupervisor = CustomUser.objects.filter(id= user.id).update(created_by_id = request.user.id)
                return Response({
                    "status": "success",
                    "message": "Successfully Inserted.",
                    "data": data,
                })
            else:
                key, value = list(serializer.errors.items())[0]
                # print(key , value)
                error_message = key + " ,"  + value[0]
                return Response({
                    "status": "error",
                    "message": error_message,

                }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as ex:
            return Response({
                "status": "error",
                "message": "Error in Field " + str(ex),
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GetDeactivatedUserList(generics.ListAPIView):
    permission_classes = [IsAuthenticated , IsAdmin]
    pagination_class = LimitOffsetPagination
    serializer_class = GetDeactivatedUserListSerializer
    model = serializer_class.Meta.model
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self):
        ward_name = self.kwargs.get('ward_name')
        group = self.kwargs.get('group')

        queryset = self.model.objects.filter( is_active=False  , created_by__groups__name = 'MOH' ,
                                            userSections__healthPost__ward__wardName = ward_name , groups__name = group ).order_by("-created_date")

        search_terms = self.request.query_params.get('search', None )
        if search_terms:
            queryset = queryset.filter(
                Q(name__icontains=search_terms) |
                Q(username__icontains=search_terms) |
                Q(phoneNumber__icontains=search_terms) |
                Q(health_Post__healthPostName__icontains=search_terms) |
                Q(userSections__healthPost__healthPostName__icontains=search_terms) )

        return queryset

    def get(self, request, *args, **kwargs):

        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer.data})

        serializer = self.get_serializer(queryset, many=True)
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                        'data': serializer.data})

class InsertUsersByMOH(generics.GenericAPIView):
    # permission_classes = [permissions.IsAuthenticated,]
    serializer_class = AddUserByMOHSerializer
    # parser_classes = [MultiPartParser]
    permission_classes = (IsAuthenticated , IsMOH)


    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)

        try:
            if serializer.is_valid():
                group = Group.objects.get(name=serializer.validated_data.get("group"))

                user = serializer.save()
                customuser = serializer.validated_data
                data = RegisterSerializer(customuser, context=self.get_serializer_context()).data
                user.groups.add(group)

                addSupervisor = CustomUser.objects.filter(id= user.id).update(created_by_id = request.user.id)
                return Response({
                    "status": "success",
                    "message": "Successfully Inserted.",
                    "data": data,
                })
            else:
                key, value = list(serializer.errors.items())[0]
                error_message = value[0]
                return Response({
                    "status": "error",
                    "message": error_message,

                }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as ex:
            return Response({
                "status": "error",
                "message": "Error in Field " + str(ex),
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class UpdateUserDetails(generics.GenericAPIView):
    serializer_class  = UpdateUsersDetailsSerializer
    permission_classes = (IsAuthenticated , IsAdmin | IsSupervisor | IsMOH)
    # parser_classes = [MultiPartParser]


    def patch(self, request, pk):
        # print(request.data)
        try:
            instance = CustomUser.objects.get(pk=pk)
        except:
            return Response({'status': 'error',
                            'message': 'deatils not found'}, status=400)

        serializer = self.get_serializer(instance , data = request.data , partial = True )
        # print(serializer)
        if serializer.is_valid():
            if "newpassword" in  serializer.validated_data:
                instance.set_password(serializer.validated_data.get("newpassword"))
                instance.save()
            serializer.save()
            return Response({"status" : "success" ,
                            "message" : "User details updated successfully"
                    },status=status.HTTP_201_CREATED)
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = value[0]
            return Response({'message': error_message,
                            'status' : 'error'}, status=400)

class deleteUser(generics.GenericAPIView):
    serializer_class = DeleteUserSerializer
    permission_classes = (IsAuthenticated , IsAdmin | IsSupervisor)

    def delete(self, request, id , *args, **kwargs):
        try:
            instance = CustomUser.objects.get(pk=id).delete()
            return Response({'status': 'success', 'message' : 'user deleted'}, status= status.HTTP_200_OK)
        except:
            return Response({'status': 'error',
                            'message': 'deatils not found'}, status=400)

class AdminChangePasswordView(generics.UpdateAPIView):
    """
    An endpoint for changing password.
    """
    serializer_class = ChangePasswordSerializer
    # model = CustomUser
    # permission_classes = (IsAuthenticated)
    def get_object(self, queryset=None):
        id = self.kwargs.get('id')
        try:
            obj = CustomUser.objects.get(id = id)
        except:
            return Response({'status': 'error',
                'message': 'user details not found'
            }, status= status.HTTP_400_BAD_REQUEST)
        return obj

    def update(self , request, *args, **kwargs):
        self.object = self.get_object()
        serializer = self.get_serializer(data=request.data)

        if serializer.is_valid():
            # Check old password
            # set_password also hashes the password that the user will get
            self.object.set_password(serializer.data.get("newpassword"))
            self.object.save()

            # sendOtp.objects.filter(registerUser_id = self.request.user.id).delete()
            response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'Password updated successfully'
            }
            return Response(response)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class userListAPI(generics.ListAPIView):
    pagination_class = LimitOffsetPagination
    serializer_class = CustomUserSerializer
    model = serializer_class.Meta.model
    # permission_classes = (IsAuthenticated , IsAdmin)
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self ):
        """
        The function returns a queryset of all objects ordered by their created date in descending order.
        """
        group = self.kwargs.get('group')

        ward_name = self.kwargs.get('ward_name')
        # section = self.model.objects.filter(groups__name = group, userSections__id = 239)
        # for i  in section:
        #     print(i.name)
        if group == 'mo':
            queryset = self.model.objects.filter(groups__name = group , dispensary__ward__wardName = ward_name ).order_by("-created_date")
        else:
            # queryset = self.model.objects.filter(groups__name = group , section__healthPost__ward__wardName = ward_name).order_by("-created_date")
            queryset = self.model.objects.filter(groups__name = group , userSections__healthPost__ward__wardName = ward_name).order_by("-created_date").distinct()

        search_terms = self.request.query_params.get('search', None )
        if search_terms:
            queryset = queryset.filter(
                Q(name__icontains=search_terms) |
                Q(username__icontains=search_terms) |
                Q(phoneNumber__icontains=search_terms) |
                Q(health_Post__healthPostName__icontains=search_terms) |
                Q(section__healthPost__healthPostName__icontains=search_terms) )

        return queryset

    def get(self, request, *args, **kwargs):

        queryset = self.get_queryset( )
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer.data})

        serializer = self.get_serializer(queryset, many=True)
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                        'data': serializer.data})

class GetWardWiseSUerList(generics.ListAPIView):
    permission_classes = [IsAuthenticated ,IsMOH]
    pagination_class = LimitOffsetPagination
    serializer_class = CustomUserSerializer
    model = serializer_class.Meta.model
    filter_backends = (filters.SearchFilter,)

    def get_queryset(self):
        """
        The function returns a queryset of all objects ordered by their created date in descending order.
        """
        group = self.kwargs.get('group')
        # wardName = self.kwargs.get('ward')
        # print(group , wardName)
        ward_id= self.request.user.ward.id
        queryset = self.model.objects.filter(groups__name = group  , userSections__healthPost__ward__id = ward_id).order_by("-created_date").distinct()

        search_terms = self.request.query_params.get('search', None)
        if search_terms:
            queryset = queryset.filter(Q(section__healthPost__ward__wardName__icontains=search_terms)|
                                        Q(userSections__icontains=search_terms) |
                                        Q(phoneNumber__icontains=search_terms) |
                                        Q(health_Post__healthPostName__icontains=search_terms))

        return queryset

    def get(self, request, *args, **kwargs):

        queryset = self.get_queryset()

        page = self.paginate_queryset(queryset)

        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'status': 'success',
                                                'message': 'Data fetched successfully',
                                                'data': serializer.data})

        serializer = self.get_serializer(queryset, many=True)
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                        'data': serializer.data})

class DownloadHealthpostwiseUserList(generics.GenericAPIView):
    # permission_classes = [IsAuthenticated , IsAdmin | IsSupervisor ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request, id, *args, **kwargs):
        try:
            healthpost = healthPost.objects.get(pk=id)
        except healthPost.DoesNotExist:
            return Response({
                "message":"No Health post exists with ID %d"%(id),
                "status":"error"
            } , status=400)
        healthpost_related_user = familyMembers.objects.filter(area__healthPost=healthpost)
        today = datetime.today().strftime('%d-%m-%Y')
        healthpost_name = healthpost.healthPostName

        if not healthpost_related_user:
            return Response({
                "message":"No data found for healthpost %s"%(healthpost_name),
                "status":"error"
            } , status=400)

        familyMember = healthpost_related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', "Address" ,'Aadhar Card', 'Abha ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  healthpost_related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                                family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]
            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(healthpost_name+"_data_"+today)
        wb.save(response)
        return response

class DownloadWardwiseUserList(generics.GenericAPIView):
    # permission_classes = [IsAuthenticated , IsAdmin | IsSupervisor ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request, id, *args, **kwargs):
        try:
            ward_obj = ward.objects.get(pk=id)
        except ward.DoesNotExist:
            return Response({
                "message":"No Ward exists with ID %d"%(id),
                "status":"error"
            } , status=400)
        # ward_related_user = familyMembers.objects.filter(familySurveyor__userSections__healthPost__ward=ward_obj)
        ward_related_user = familyMembers.objects.filter(area__healthPost__ward=ward_obj)
        today = datetime.today().strftime('%d-%m-%Y')
        ward_name = ward_obj.wardName

        if not ward_related_user:
            return Response({
                "message":"No data found for ward %s"%(ward_name),
                "status":"error"
            } , status= 400)

        familyMember = ward_related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', "Address" ,'Aadhar Card', 'Abha ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  ward_related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                               family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]


            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format("Ward_"+ward_name+"_data_"+today)
        wb.save(response)
        return response

class DownloadAllWardUserList(generics.GenericAPIView):
    # permission_classes = [IsAuthenticated , IsAdmin | IsSupervisor ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request, *args, **kwargs):

        ward_related_user = familyMembers.objects.all()
        today = datetime.today().strftime('%d-%m-%Y')

        familyMember = ward_related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', 'Address', 'Aadhar Card', 'Abha ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  ward_related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address ,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y %I:%M:%S %p'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                                family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]
            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format("All_Ward_data_"+today)
        wb.save(response)
        return response

class DownloadDispensarywiseUserList(generics.GenericAPIView):

    # permission_classes = [IsAuthenticated , IsMO ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():

                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request ,id, *args, **kwargs):
        try:
            dispensary_obj = dispensary.objects.get(pk=id)
        except dispensary.DoesNotExist:
            return Response({
                "message":"No Dispensary exists with ID %d"%(id),
                "status":"error"
            } , status=400)
        dispensary_related_user = familyMembers.objects.filter(area__dispensary=dispensary_obj)
        today = datetime.today().strftime('%d-%m-%Y')
        dispensary_name = dispensary_obj.dispensaryName

        if not dispensary_related_user:
            return Response({
                "message":"No data found for dispensary %s"%(dispensary_name),
                "status":"error"
            } , status=400)

        familyMember = dispensary_related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', 'Address', 'Aadhar Card', 'Abha ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  dispensary_related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address ,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                               family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]
            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(dispensary_name+"_data_"+today)
        wb.save(response)
        return response

""" MOH API's """

class MOHDashboardView(generics.GenericAPIView): #Modified
    permission_classes= (IsAuthenticated , IsMOH)
    queryset = familyMembers.objects.all()
    FamilySurvey_count = familyHeadDetails.objects.all()
    CustomUser_queryset = CustomUser.objects.all()

    def get(self, request ,  *args, **kwargs):

        ward_id = request.user.ward_id
        healthpost_id = request.query_params.get('healthpost_id', None)

        # Staff statistics.
        CHV_ASHA_count = self.CustomUser_queryset.filter(userSections__healthPost__ward__id = ward_id ,groups__name='CHV-ASHA').distinct().count()
        MO_count = self.CustomUser_queryset.filter(dispensary__ward__id = ward_id, groups__name='mo').count()
        ANM_count = self.CustomUser_queryset.filter(userSections__healthPost__ward__id = ward_id ,groups__name='healthworker').distinct().count()

        today = timezone.now().date()
        # Get the date 7 days ago
        seven_days_ago = timezone.now() - timedelta(days=7)
        # Get users who have logged in within the last 7 days
        active_ANMs = AuthToken.objects.filter(created__gte=seven_days_ago,
                                                user__userSections__healthPost__ward__id=ward_id,
                                                user__groups__name='healthworker').distinct().count()
        inactive_ANMs = ANM_count - active_ANMs

        if healthpost_id:

            # Distinct and common queries of survey data
            healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost_id)

            # Healthppost related survey data
            aggregated_data = get_aggregated_data(healthpost_queryset, familySurvey_queryset)

        else:
            # Distinct and common queries of survey data
            ward_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward__id=ward_id).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__ward__id=ward_id)

            # Ward related survey data
            aggregated_data = get_aggregated_data(ward_queryset, familySurvey_queryset)

        return Response({
            'active_ANMs': active_ANMs,
            'inactive_ANMs': inactive_ANMs,
            'CHV_ASHA_count' : CHV_ASHA_count,
            'MO_count' : MO_count,
            'ANM_count' : ANM_count,
            **aggregated_data}, status=status.HTTP_200_OK)

class MOHDashboardExportView(generics.GenericAPIView): #Modified

    permission_classes = (IsAuthenticated, IsMOH)
    FamilySurvey_count = familyHeadDetails.objects.all()
    CustomUser_queryset = CustomUser.objects.all()

    def add_headers(self, sheet, *args):
        """This function takes the sheet and add merged headers to that sheet"""
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def get_queryset(self):
        """
        The function returns a queryset of all objects ordered by their created date in descending order.
        """
        queryset = familyMembers.objects.all()

        return queryset

    def get(self, request, *args, **kwargs):

        healthpost_id = request.query_params.get('healthpost_id', None)
        ward_id = request.user.ward_id
        group_name = request.query_params.get('group_name', None)

        today = datetime.today().strftime('%d-%m-%Y')

        # Name of the excel file to be exported.
        excel_name = ""

        if not group_name:
            return Response({
                    "message":"Please provide group name",
                    "status":"error"
                }, status=400)

        group_name = group_name.strip()

        if group_name == "healthworker":
            data_list = [['Ward Name', 'Health Post Name', 'ANM/Co-ordinator', 'ANM Mobile number', 'Families Enrolled',
                          'Citizens Enrolled', 'CBAC Filled', 'Citizens 60 years + enrolled',
                          'Citizens 30 years + enrolled', 'Males Enrolled', 'Females Enrolled',
                          'Transgender Enrolled', 'ABHA ID Generated', 'Vulnerable Citizen', 'Diabetes',
                          'Hypertension', 'Oral Cancer', 'Cervical cancer', 'COPD', 'Eye Disorder',
                          'ENT Disorder', 'Asthma', 'Alzheimers', 'TB', 'leprosy','Breast cancer', 'Other Communicable',
                          'Blood collected at home', 'blood collected at center', 'Blood Collection Denied By AMO',
                          'Blood Collection Denied By Citizen', 'Total Reports Generated', 'Tests Assigned',
                          'Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment',
                          'Referral to HBT polyclinic for Physician consultation',
                          'Referral to Peripheral Hospital / Special Hospital for management of Complication',
                          'Referral to Medical College for management of Complication',
                          'Referral to Private facility']]
        elif group_name == "CHV-ASHA":
            data_list = [['Ward Name', 'Health Post Name', 'CHV-ASHA', 'CHV-ASHA Mobile number','Families Enrolled',
                          'Citizens Enrolled', 'CBAC Filled', 'Citizens 60 years + enrolled',
                          'Citizens 30 years + enrolled', 'Males Enrolled', 'Females Enrolled',
                          'Transgender Enrolled', 'ABHA ID Generated', 'Vulnerable Citizen', 'Diabetes',
                          'Hypertension', 'Oral Cancer', 'Cervical cancer', 'COPD', 'Eye Disorder',
                          'ENT Disorder', 'Asthma', 'Alzheimers', 'TB', 'leprosy','Breast cancer', 'Other Communicable',
                          'Blood collected at home', 'blood collected at center', 'Blood Collection Denied By AMO',
                          'Blood Collection Denied By Citizen', 'Total Reports Generated', 'Tests Assigned',
                          'Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment',
                          'Referral to HBT polyclinic for Physician consultation',
                          'Referral to Peripheral Hospital / Special Hospital for management of Complication',
                          'Referral to Medical College for management of Complication',
                          'Referral to Private facility']]

        header1 = {'Citizen Details':len(data_list[0][0:14]), 'Dieases Suspected' : len(data_list[0][14:27]),
                   'Blood Collection' : len(data_list[0][27:33]), 'Referrals' : len(data_list[0][33:])}

        if healthpost_id:
            try:
                healthpost = healthPost.objects.get(pk=healthpost_id)
            except healthPost.DoesNotExist:
                return Response({
                    "message":"No Health post exists with ID %d"%(id),
                    "status":"error"
                } , status= 400 )

            healthpost_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()
            healthpost_name = healthpost.healthPostName

            if not healthpost_related_user:
                return Response({
                    "message":"No data found for healthpost %s"%(healthpost_name),
                    "status":"error"} , status = 400)

            excel_name = "healthpost_" + healthpost_name + "_staff_summary_"

            health_workers = CustomUser.objects.filter(groups__name = group_name , userSections__healthPost__id = healthpost_id).distinct().order_by("name")

            for healthworker in health_workers:
                # Healthpost and Health Worker data
                healthpost_data = [healthpost.ward.wardName, healthpost.healthPostName,
                                   healthworker.name, healthworker.phoneNumber]

                # Distinct and common queries of survey data
                healthpost_queryset = healthpost_related_user.filter(familySurveyor__id=healthworker.id)
                familySurvey_queryset = self.FamilySurvey_count.filter(user__id=healthworker.id)

                # Healthppost related survey data
                aggregated_data = get_aggregated_data(healthpost_queryset, familySurvey_queryset, aggregate_type="export")

                combined_survey_data = healthpost_data + aggregated_data

                data_list.append(combined_survey_data)
        else:
            try:
                ward_obj = ward.objects.get(pk=ward_id)
            except ward.DoesNotExist:
                return Response({
                    "message":"No ward exists with ID %d"%(ward_id),
                    "status":"error"
                } , status = 400)

            ward_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward__id=ward_id).exists()
            ward_name = ward_obj.wardName

            if not ward_related_user:
                return Response({
                    "message":"No data found for ward %s"%(ward_name),
                    "status":"error"
                } , status = 400)

            excel_name = "ward_" + ward_name + "_staff_summary_"

            healthposts = healthPost.objects.filter(ward__id = ward_id)

            for healthpost in healthposts:

                health_workers = CustomUser.objects.filter(groups__name = group_name , userSections__healthPost__id = healthpost.id).distinct().order_by("name")

                for healthworker in health_workers:
                    # Ward and Health Worker data
                    ward_data = [ward_name, healthpost.healthPostName,
                                 healthworker.name, healthworker.phoneNumber]

                    # Common querysets of survey
                    ward_queryset = self.get_queryset().filter(familySurveyor__id=healthworker.id)
                    familySurvey_queryset = self.FamilySurvey_count.filter(user__id=healthworker.id)

                    # Ward related survey data
                    aggregated_data = get_aggregated_data(ward_queryset, familySurvey_queryset, aggregate_type="export")

                    combined_survey_data = ward_data + aggregated_data

                    data_list.append(combined_survey_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(excel_name + today)
        wb.save(response)
        return response

class MOHDashboardTabView(generics.GenericAPIView): #Modified

    permission_classes = (IsAuthenticated, IsMOH)
    queryset = familyMembers.objects.all()
    FamilySurvey_count = familyHeadDetails.objects.all()
    pagination_class = CustomPageNumberPagination  # Add pagination class

    def get(self, request):

        healthpost_id = request.query_params.get('healthpost_id', None)
        ward_id = request.user.ward_id

        data_list =[]

        if healthpost_id:
            try:
                healthpost = healthPost.objects.get(pk=healthpost_id)
            except healthPost.DoesNotExist:
                return Response({
                    "message":"No Health post exists for id %s"%(healthpost_id),
                    "status":"error"
                } , status= 400 )

            healthpost_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()

            # Healthpost Survey data
            healthpost_data = {"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName}

            if not healthpost_related_user:
                data_list.append({**healthpost_data,
                "total_family_count":0 , "total_citizen_count":0 ,
                "total_cbac_count":0 , "citizen_above_60":0 ,
                "citizen_above_30":0 , "male":0 , "female":0 ,"transgender":0 ,"total_AbhaCreated":0 ,
                "total_diabetes":0 , "hypertension":0 , "total_oral_cancer":0 , "total_cervical_cancer":0 , "total_COPD_count":0 , "total_eye_problem":0 ,
                "total_ent_disorder":0 , "total_asthma":0, "total_Alzheimers":0, "total_tb_count":0 ,"total_leprosy":0 , "total_breast_cancer":0 , "total_communicable":0 ,
                "blood_collected_home":0 , "blood_collected_center":0 ,  "denied_by_mo_count":0  ,  "denied_by_mo_individual":0 , "TestReportGenerated":0 , "total_LabTestAdded":0,
                "Referral_choice_Referral_to_Mun_Dispensary":0 ,
                "Referral_choice_Referral_to_HBT_polyclinic":0,
                "Referral_choice_Referral_to_Peripheral_Hospital":0,
                "Referral_choice_Referral_to_Medical_College":0 ,
                "Referral_choice_Referral_to_Private_facility":0,
                "total_vulnerable":0})

                return Response({
                    "message":"Data fetched successfully",
                    "status":"success","results":data_list} , status=status.HTTP_200_OK)

            # Common querysets of survey
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost_id)

            # Healthppost related survey data
            aggregated_data = get_aggregated_data(healthpost_related_user, familySurvey_queryset, aggregate_type="tab")

            combined_survey_data = {**healthpost_data, **aggregated_data}

            data_list.append(combined_survey_data)
        else:
            healthposts = healthPost.objects.filter(ward__id=ward_id)

            for healthpost in healthposts:
                # Healthpost Survey data
                healthpost_data = {"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName}

                # Common querysets of survey
                healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost.id).distinct()
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost.id)

                # Healthppost related survey data
                aggregated_data = get_aggregated_data(healthpost_queryset, familySurvey_queryset, aggregate_type="tab")

                combined_survey_data = {**healthpost_data, **aggregated_data}

                data_list.append(combined_survey_data)

        page = self.paginate_queryset(data_list)
        if page is not None:
            return self.get_paginated_response(page)
        return Response({
                "message":"Data fetched successfully",
                "status":"success",
                "data":data_list}, status=status.HTTP_200_OK)

""" Admin API's """

class AdminDashboardView(generics.GenericAPIView): #Modified
    permission_classes= (IsAuthenticated, IsAdmin|IsViewAdmin)
    queryset = familyMembers.objects.all()
    FamilySurvey_count = familyHeadDetails.objects.all()
    CustomUser_queryset = CustomUser.objects.all()

    def get(self, request ,  *args, **kwargs):

        healthpost_id = request.query_params.get('healthpost_id', None)
        ward_id = request.query_params.get('wardId', None)

        # Staff Counts
        CHV_ASHA_count = self.CustomUser_queryset.filter(groups__name='CHV-ASHA').count()
        MO_count = self.CustomUser_queryset.filter(groups__name='mo').count()
        ANM_count = self.CustomUser_queryset.filter(groups__name='healthworker').count()

        today = timezone.now().date()

        # Get the date 7 days ago
        seven_days_ago = timezone.now() - timedelta(days=7)

        # Get users who have logged in within the last 7 days
        active_ANMs = AuthToken.objects.filter(created__gte=seven_days_ago, user__groups__name='healthworker').distinct().count()
        inactive_ANMs = ANM_count - active_ANMs

        if healthpost_id:
            # Distinct and common queries of survey data
            healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost_id)

            # Healthppost related survey data
            aggregated_data = get_aggregated_data(healthpost_queryset, familySurvey_queryset)
        elif ward_id :
            # Distinct and common queries of survey data
            ward_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward__id=ward_id).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__ward__id=ward_id)

            # Ward related survey data
            aggregated_data = get_aggregated_data(ward_queryset, familySurvey_queryset)
        else:
            # Survey data
            aggregated_data = get_aggregated_data(self.get_queryset(), self.FamilySurvey_count)

        return Response({
            'active_ANMs': active_ANMs,
            'inactive_ANMs': inactive_ANMs,
            'CHV_ASHA_count' : CHV_ASHA_count,
            'MO_count' : MO_count,
            'ANM_count' : ANM_count,
            **aggregated_data}, status=status.HTTP_200_OK)

class AdminDashboardExportView(generics.GenericAPIView): #Modified
    permission_classes= (IsAuthenticated , IsAdmin|IsViewAdmin)
    FamilySurvey_count = familyHeadDetails.objects.all()
    queryset = familyMembers.objects.all()

    def add_headers(self, sheet, *args):
        """This function takes the sheet and add merged headers to that sheet"""
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def get(self, request):
        # aggregate query param to determine whether to return all data or partial data.
        aggregate = request.query_params.get('aggregate', 'false').lower() in ['true', 'yes']

        ward_id = request.query_params.get('ward_id', None)
        healthpost_id = request.query_params.get('healthpost_id', None)

        # filter_by query param to filter the aggregated counts either by "ward" or "healthpost".
        filter_by = request.query_params.get('filter_by', None)

        today = datetime.today().strftime('%d-%m-%Y')
        # Name of the excel file to be exported.
        excel_name = ""

        if filter_by == "healthpost":
            excel_data = [['Ward Name', 'Health Post Name', 'Families Enrolled', 'Citizens Enrolled', 'CBAC Filled',
                        'Citizens 60 years + enrolled', 'Citizens 30 years + enrolled', 'Males Enrolled',
                        'Females Enrolled', 'Transgender Enrolled', 'ABHA ID Generated', 'Vulnerable Citizen',
                        'Diabetes', 'Hypertension', 'Oral Cancer', 'Cervical cancer', 'COPD', 'Eye Disorder',
                        'ENT Disorder', 'Asthma', 'Alzheimers', 'TB', 'Leprosy', 'Breast cancer', 'Other Communicable',
                        'Blood collected at home', 'blood collected at center', 'Blood Collection Denied By AMO',
                        'Blood Collection Denied By Citizen', 'Total Reports Generated', 'Tests Assigned',
                        'Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment',
                        'Referral to HBT polyclinic for Physician consultation',
                        'Referral to Peripheral Hospital / Special Hospital for management of Complication',
                        'Referral to Medical College for management of Complication',
                        'Referral to Private facility']]

            header1 = {'Citizen Details':len(excel_data[0][0:12]),'Dieases Suspected': len(excel_data[0][12:25]),
                    'Blood Collection' : len(excel_data[0][25:31]), 'Referrals' : len(excel_data[0][31:])}
        else:
            excel_data = [['Ward Name', 'Families Enrolled', 'Citizens Enrolled', 'CBAC Filled',
                        'Citizens 60 years + enrolled', 'Citizens 30 years + enrolled', 'Males Enrolled',
                        'Females Enrolled', 'Transgender Enrolled', 'ABHA ID Generated', 'Vulnerable Citizen',
                        'Diabetes', 'Hypertension', 'Oral Cancer', 'Cervical cancer', 'COPD', 'Eye Disorder',
                        'ENT Disorder', 'Asthma', 'Alzheimers', 'TB', 'Leprosy', 'Breast cancer', 'Other Communicable',
                        'Blood collected at home', 'blood collected at center', 'Blood Collection Denied By AMO',
                        'Blood Collection Denied By Citizen', 'Total Reports Generated', 'Tests Assigned',
                        'Referral to Mun. Dispensary / HBT for Blood Test / Confirmation / Treatment',
                        'Referral to HBT polyclinic for Physician consultation',
                        'Referral to Peripheral Hospital / Special Hospital for management of Complication',
                        'Referral to Medical College for management of Complication',
                        'Referral to Private facility']]

            header1 = {'Citizen Details':len(excel_data[0][0:11]),'Dieases Suspected': len(excel_data[0][11:24]),
                    'Blood Collection' : len(excel_data[0][24:30]), 'Referrals' : len(excel_data[0][30:])}

        if filter_by == "healthpost":
            if healthpost_id:
                try:
                    healthpost = healthPost.objects.get(pk=healthpost_id)
                except healthPost.DoesNotExist:
                    return Response({
                        "message":"No Health post exists with ID %d"%(id),
                        "status":"error"
                    }, status=status.HTTP_400_BAD_REQUEST)

                healthpost_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost_id).distinct()
                healthpost_name = healthpost.healthPostName

                if not healthpost_related_user:
                    return Response({
                        "message":"No data found for healthpost %s"%(healthpost_name),
                        "status":"error"
                    }, status=status.HTTP_400_BAD_REQUEST)

                excel_name = "healthpost_" + healthpost_name + "_summary_"

                # Distinct and common queries of survey data
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost_id)

                # healthpost related survey data
                aggregated_data = get_aggregated_data(healthpost_related_user, familySurvey_queryset, aggregate_type="export")

                export_data = [healthpost.ward.wardName, healthpost_name] + aggregated_data
                excel_data.append(export_data)
            elif ward_id:
                try:
                    ward_obj = ward.objects.get(pk=ward_id)
                except ward.DoesNotExist:
                    return Response({
                        "message":"No ward exists with ID %s"%(ward_id),
                        "status":"error"
                    }, status=status.HTTP_400_BAD_REQUEST)

                ward_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward__id=ward_id).exists()
                ward_name = ward_obj.wardName

                if not ward_related_user:
                    return Response({
                        "message":"No data found for ward %s"%(ward_name),
                        "status":"error"
                    }, status=status.HTTP_400_BAD_REQUEST)

                excel_name = "ward_" + ward_name + "_healthposts_summary_"

                healthposts = healthPost.objects.filter(ward_id=ward_id)

                for healthpost in healthposts:
                    # Common querysets of survey
                    healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost_id=healthpost.id).distinct()
                    familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost_id=healthpost.id)

                    # healthpost related survey data
                    aggregated_data = get_aggregated_data(healthpost_queryset, familySurvey_queryset, aggregate_type="export")

                    export_data = [ward_name, healthpost.healthPostName] + aggregated_data
                    excel_data.append(export_data)
            elif aggregate:
                excel_name = "healthposts_summary_"
                healthposts = healthPost.objects.all().order_by('ward__wardName')
                for healthpost in healthposts:
                    # Distinct and common queries of survey data
                    healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost_id=healthpost.id).distinct()
                    familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost_id=healthpost.id)

                    # healthpost related survey data
                    aggregated_data = get_aggregated_data(healthpost_queryset, familySurvey_queryset, aggregate_type="export")

                    export_data = [healthpost.ward.wardName, healthpost.healthPostName] + aggregated_data
                    excel_data.append(export_data)
            else:
                return Response({
                    "message":"Please Provide either aggregate, ward_id or healthpost_id.",
                    "status":"error"
                }, status=status.HTTP_400_BAD_REQUEST)
        elif filter_by == "ward":
            if ward_id:
                try:
                    ward_obj = ward.objects.get(pk=ward_id)
                except ward.DoesNotExist:
                    return Response({
                        "message":"No ward exists with ID %s"%(ward_id),
                        "status":"error"
                    }, status=status.HTTP_400_BAD_REQUEST)

                ward_related_user = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward_id=ward_id).distinct()
                ward_name = ward_obj.wardName

                if not ward_related_user:
                    return Response({
                        "message":"No data found for ward %s"%(ward_name),
                        "status":"error"
                    }, status=status.HTTP_400_BAD_REQUEST)

                excel_name = "ward_" + ward_name + "_summary_"

                # Distinct and common queries of survey data
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__ward_id=ward_id)

                # Ward related survey data
                aggregated_data = get_aggregated_data(ward_related_user, familySurvey_queryset, aggregate_type="export")

                export_data = [ward_name] + aggregated_data
                excel_data.append(export_data)
            elif aggregate:
                excel_name = "wards_summary_"
                wards = ward.objects.all().order_by("wardName")
                for ward_obj in wards:
                    # Distinct and common queries of survey data
                    ward_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__ward_id=ward_obj.id).distinct()
                    familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__ward_id=ward_obj.id)

                    # Ward related survey data
                    aggregated_data = get_aggregated_data(ward_queryset, familySurvey_queryset, aggregate_type="export")

                    export_data = [ward_obj.wardName] + aggregated_data
                    excel_data.append(export_data)
            else:
                return Response({
                    "message":"Please Provide either aggregate, ward_id or healthpost_id.",
                    "status":"error"
                }, status=status.HTTP_400_BAD_REQUEST)
        elif aggregate:
            excel_name = "dashboard_data_summary_"
            # Aggregated survey data
            aggregated_data = get_aggregated_data(self.get_queryset(), self.FamilySurvey_count, aggregate_type="export")

            export_data = ["All"] + aggregated_data
            excel_data.append(export_data)
        else:
            return Response({
                "message":"Please Provide either aggregate or filter_by.",
                "status":"error"
            }, status=status.HTTP_400_BAD_REQUEST)

        workbook = openpyxl.Workbook()
        working_sheet = workbook.active
        self.add_headers(working_sheet, header1)
        for row in excel_data:
            working_sheet.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(excel_name + today)
        workbook.save(response)
        return response

class AdminDashboardTabView(generics.GenericAPIView): #Modified
    permission_classes = (IsAuthenticated , IsAdmin|IsViewAdmin)
    FamilySurvey_count = familyHeadDetails.objects.all().order_by('user__userSections__healthPost__ward')
    queryset = familyMembers.objects.all()
    healthposts = healthPost.objects.all().order_by('ward__wardName')
    pagination_class = CustomPageNumberPagination  # Add pagination class

    def get(self, request):

        healthpost_id = request.query_params.get('healthpost_id', None)
        ward_id = request.query_params.get('wardId', None)

        data_list = [] # List for storing survey data
        today = datetime.today().strftime('%d-%m-%Y')

        if ward_id and not healthpost_id:
            try:
                ward_name = ward.objects.get(pk=ward_id)
            except ward.DoesNotExist:
                return Response({
                    "message":"No ward exists with ID %s"%(ward_id),
                    "status":"error"
                } , status = 400)

            ward_related_user = familyMembers.objects.filter(familySurveyor__userSections__healthPost__ward__id=ward_id).exists()

            if not ward_related_user:
                data_list.append({"wardName":ward_name.wardName, "healthPostName":"",
                "total_family_count":0, "total_citizen_count":0, "total_cbac_count":0, "citizen_above_60":0,
                "citizen_above_30":0, "male":0, "female":0, "transgender":0, "total_AbhaCreated":0,
                "total_diabetes":0, "hypertension":0, "total_oral_cancer":0, "total_cervical_cancer":0,
                "total_COPD_count":0, "total_eye_problem":0, "total_ent_disorder":0, "total_asthma":0,
                "total_Alzheimers":0, "total_tb_count":0,"total_leprosy":0, "total_breast_cancer":0,
                "total_communicable":0, "blood_collected_home":0, "blood_collected_center":0,
                "denied_by_mo_count":0 , "denied_by_mo_individual":0, "TestReportGenerated":0,
                "total_LabTestAdded":0, "Referral_choice_Referral_to_Mun_Dispensary":0,
                "Referral_choice_Referral_to_HBT_polyclinic":0, "Referral_choice_Referral_to_Peripheral_Hospital":0,
                "Referral_choice_Referral_to_Medical_College":0, "Referral_choice_Referral_to_Private_facility":0,
                "total_vulnerable":0})

                return Response({
                    "message":"Data fetched successfully",
                    "status":"success",
                    "results":data_list}, status=status.HTTP_200_OK)

            healthposts = self.healthposts.filter(ward__id=ward_id)

            for healthpost in healthposts:
                # Healthpost Survey data
                healthpost_data = {"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName}

                # Common querysets of survey
                healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost__id=healthpost.id).distinct()
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost__id=healthpost.id)

                # Healthppost related survey data
                aggregated_data = get_aggregated_data(healthpost_queryset, familySurvey_queryset, aggregate_type="tab")

                combined_survey_data = {**healthpost_data, **aggregated_data}

                data_list.append(combined_survey_data)
        elif ward_id and healthpost_id:
            try:
                healthpost = healthPost.objects.get(pk=healthpost_id, ward_id=ward_id)
            except healthPost.DoesNotExist:
                return Response({
                    "message":"No Health post exists with ID %s"%(healthpost_id),
                    "status":"error",
                } , status = 400)

            healthpost_related_user = familyMembers.objects.filter(familySurveyor__userSections__healthPost_id=healthpost_id).distinct()
            healthpost_name = healthpost.healthPostName

            # Healthpost Survey data
            healthpost_data = {"wardName":healthpost.ward.wardName, "healthPostName":healthpost_name}

            if not healthpost_related_user:
                data_list.append({**healthpost_data,
                "total_family_count":0, "total_citizen_count":0, "total_cbac_count":0, "citizen_above_60":0,
                "citizen_above_30":0, "male":0, "female":0, "transgender":0, "total_AbhaCreated":0,
                "total_diabetes":0, "hypertension":0, "total_oral_cancer":0, "total_cervical_cancer":0,
                "total_COPD_count":0, "total_eye_problem":0, "total_ent_disorder":0, "total_asthma":0,
                "total_Alzheimers":0, "total_tb_count":0,"total_leprosy":0, "total_breast_cancer":0,
                "total_communicable":0, "blood_collected_home":0, "blood_collected_center":0,
                "denied_by_mo_count":0 , "denied_by_mo_individual":0, "TestReportGenerated":0,
                "total_LabTestAdded":0, "Referral_choice_Referral_to_Mun_Dispensary":0,
                "Referral_choice_Referral_to_HBT_polyclinic":0, "Referral_choice_Referral_to_Peripheral_Hospital":0,
                "Referral_choice_Referral_to_Medical_College":0, "Referral_choice_Referral_to_Private_facility":0,
                "total_vulnerable":0})

                return Response({
                    "message":"Data fetched successfully",
                    "status":"success",
                    "results":data_list}, status=status.HTTP_200_OK)

            # Common querysets of survey
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost_id=healthpost_id)

            # Healthppost related survey data
            aggregated_data = get_aggregated_data(healthpost_related_user, familySurvey_queryset, aggregate_type="tab")

            combined_survey_data = {**healthpost_data, **aggregated_data}

            data_list.append(combined_survey_data)
        else:
            for healthpost in self.healthposts:
                # Healthpost Survey data
                healthpost_data = {"wardName":healthpost.ward.wardName, "healthPostName":healthpost.healthPostName}

                # Distinct and common queries of survey data
                healthpost_queryset = self.get_queryset().filter(familySurveyor__userSections__healthPost_id=healthpost.id).distinct()
                familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__healthPost_id=healthpost.id)

                # Healthppost related survey data
                aggregated_data = get_aggregated_data(healthpost_queryset, familySurvey_queryset, aggregate_type="tab")

                combined_survey_data = {**healthpost_data, **aggregated_data}

                data_list.append(combined_survey_data)

        page = self.paginate_queryset(data_list)
        if page is not None:
            return self.get_paginated_response(page)

        return Response({
                "message":"Data fetched successfully",
                "status":"success",
                "data":data_list}, status=status.HTTP_200_OK)

class GetAllUserDetails(generics.GenericAPIView):
    def get(self , request ):
        data = {'wards': []}

        wards = ward.objects.all().order_by("wardName")

        for w in wards:
            wrd = {'ward': w.wardName, 'healthPosts': []}

            healthposts = healthPost.objects.filter(ward=w)
            for hp in healthposts:
                areaData = area.objects.filter(healthPost_id = hp.id).values("areas")
                areaList =  [item["areas"] for item in areaData]
                health_post_info = {'healthPost': hp.healthPostName,'areaList':areaList, 'sections': []}

                sections = section.objects.filter(healthPost=hp)
                for sec in sections:
                    section_info = {'sectionName': sec.sectionName, 'anms': [],}

                    # ANMs
                    # anms = CustomUser.objects.filter(section_id=sec.id, groups__name="healthworker")
                    anms = CustomUser.objects.filter(userSections__in=[sec], groups__name="healthworker")

                    for anm in anms:
                        anm_info = {'anmName': anm.name, 'chvs': []}

                        # CHVs under the ANM
                        chvs = CustomUser.objects.filter(ANM_id = anm.id,groups__name="CHV-ASHA")
                        for chv in chvs:
                            chv_info = {'chvName': chv.name}
                            anm_info['chvs'].append(chv_info)

                        section_info['anms'].append(anm_info)

                    health_post_info['sections'].append(section_info)

                wrd['healthPosts'].append(health_post_info)

            data['wards'].append(wrd)

        return Response({
            'status': 'success',
            'message': 'Successfully Fetched',
            'data': data,
        })