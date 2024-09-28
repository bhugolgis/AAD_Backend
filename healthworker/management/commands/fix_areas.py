from django.core.management.base import BaseCommand, CommandError
from database.models import healthPost, dispensary, ward, area
from django.db import transaction
from openpyxl import load_workbook, Workbook
import os
import re


class Command(BaseCommand):
    help = "Fixes the incorrect areas for all the healthposts and dispensary"
    error_excel_data = [] #list for storing the areas which resulted in error.

    def add_arguments(self, parser):
        parser.add_argument("--folder-path", type=str, required=True,
                            help="Path of the folder consisting of areas linkages")

    def handle(self, *args, **options):
        folder_path = options.get("folder_path")
        self.fix_area_mapping(folder_path)

    def fix_area_mapping(self, excel_folder_path:str) -> None:
        """
        Handles the entire process of fixing the area mapping from the excel files.

        Parameters:
        - excel_folder_path (str): The path to the Excel files containing area mapping data.

        Returns: None.
        """
        for file in os.listdir(excel_folder_path):
            if file.endswith(".xlsx") or file.endswith(".xls"):

                self.error_excel_data.append(["Ward", "Health Post", "Health Post ID", "Dispensary",
                                              "Dispensary ID", "Area", "Action", "Remark"])

                ward_name = file.split("-")[0].strip()
                file_path = os.path.join(excel_folder_path, file)
                excel_data = self.read_excel_data(file_path)
                if not excel_data:
                    continue
                self.update_area_mapping_database(ward_name, excel_data)
                if len(self.error_excel_data) > 1:
                    workbook = Workbook()
                    sheet = workbook.active

                    for row in self.error_excel_data:
                        sheet.append(row)
                    error_file_path = os.path.join(excel_folder_path, "Errors", ward_name + ".xlsx")
                    workbook.save(error_file_path)
                self.stdout.write(self.style.SUCCESS(f"\n------------ Fixed Areas For the Ward: {ward_name} -------------\n"))

    def read_excel_data(self, excel_file_path:str) -> list:
        """
        Reads data from an Excel file and returns it as a list of dictionaries.

        Parameters:
        - excel_file_path (str): The path to the Excel file.

        Returns: A list of dictionaries, where each dictionary represents a row of data from the Excel file.
        """
        excel_data = [] # List for storing excel data

        # Dictionary for storing the row data of excel
        row_data = {
            "serial_no":None,
            "healthpost":None,
            "dispensary":None,
            "sections":[]
        }

        sections = []
        areas = []
        duplicate_check = []

        section = {
            "section_id":None,
            "areas":[]
        }

        try:
            workbook = load_workbook(excel_file_path)
        except FileNotFoundError:
            self.stderr.write("File not found at: %s"%(excel_file_path))
            return None
        except PermissionError:
            self.stderr.write("Permission error while opening file at: %s"%(excel_file_path))
            return None

        sheet = workbook.active
        for row in sheet.iter_rows(min_row=3, min_col=1, values_only=True):
            if row[13]:
                try:
                    int(row[13][0])
                    area_name = row[13][2:].strip()
                except Exception:
                    area_name = row[13].strip()
            else:
                area_name = None
            serial_no = row[0]
            dispensary = row[2]
            healthpost = row[3]
            section_id = row[5]

            if not serial_no:
                if not section_id:
                    if area_name and area_name not in duplicate_check:
                        duplicate_check.append(area_name)
                        areas.append(area_name)
                else:
                    if section["section_id"]:
                        section["areas"] = areas[:]
                        sections.append(section.copy())
                        areas.clear()
                    section["section_id"] = section_id

                    if area_name and area_name not in duplicate_check:
                        duplicate_check.append(area_name)
                        areas.append(area_name)

            else:
                if row_data["serial_no"]:
                    row_data["sections"] = sections[:]
                    excel_data.append(row_data.copy())
                    sections.clear()
                    duplicate_check.clear()

                row_data["serial_no"] = serial_no
                row_data["healthpost"] = healthpost
                row_data["dispensary"] = re.split("disp", dispensary, flags=re.IGNORECASE)[0]

                if not section_id:
                    if area_name and area_name not in duplicate_check:
                        duplicate_check.append(area_name)
                        areas.append(area_name)
                else:
                    if section["section_id"]:
                        section["areas"] = areas[:]
                        if section["areas"]:
                            sections.append(section.copy())
                            areas.clear()
                    section["section_id"] = section_id

                    if area_name and area_name not in duplicate_check:
                        duplicate_check.append(area_name)
                        areas.append(area_name)
        else:
            row_data["sections"] = sections[:]
            excel_data.append(row_data.copy())

        return excel_data

    def update_area_mapping_database(self, ward_name:str, mapping_data:list):
        """
        Updates the area mapping in the database for the given area mapping data.

        Parameters:
        - ward_name (str): A string representing the ward name.
        - mapping_data (list): A list of dictionaries conataining the data of area mapping excel.

        Returns: None.
        """
        try:
            ward_id = ward.objects.get(wardName=ward_name).id
        except ward.DoesNotExist:
            try:
                ward_name = "/".join(list(ward_name))
                ward_id = ward.objects.get(wardName=ward_name).id
            except ward.DoesNotExist:
                self.stdout.write(self.style.WARNING("Ward %s Does not exists"%(ward_name)))
                return

        for data in mapping_data:

            healthpost_name = data["healthpost"].strip()
            dispensary_name = data["dispensary"].strip()
            sections = data["sections"]

            try:
                healthpost_id = healthPost.objects.get(healthPostName__icontains=healthpost_name, ward__id=ward_id).id
            except healthPost.DoesNotExist:
                self.stdout.write(self.style.WARNING("Health Post %s does not exist in ward %s"%(healthpost_name, ward_name)))
                self.error_excel_data.append([ward_name, healthpost_name, "NA", "NA", "NA", "NA", "NA",
                                              "Health Post %s does not exist in ward %s"%(healthpost_name, ward_name)])
                continue

            try:
                dispensary_id = dispensary.objects.get(dispensaryName__icontains=dispensary_name, ward_id=ward_id).id
            except dispensary.DoesNotExist:
                self.stdout.write(self.style.WARNING("Dispensary %s does not exist in ward %s"%(dispensary_name, ward_name)))
                self.error_excel_data.append([ward_name, "NA", "NA", dispensary_name, "NA", "NA", "NA",
                                              "Dispensary %s does not exist in ward %s"%(dispensary_name, ward_name)])
                continue

            db_areas = area.objects.filter(healthPost_id=healthpost_id, dispensary_id=dispensary_id)

            processed_areas = self.process_areas(list(db_areas), sections)

            for processed_area in processed_areas:
                area_name = processed_area["area_name"]
                action = processed_area["action"]
                similar_areas = processed_area["similar_areas"]

                if action in ['Insert', 'Update']:
                    if len(similar_areas) > 1:
                        areas_with_relations = []
                        areas_without_relations = []

                        # Categorize areas into those with and without relations
                        for area_obj in similar_areas:
                            if area_obj.familyheaddeatils_area.exists():
                                areas_with_relations.append(area_obj)
                            else:
                                areas_without_relations.append(area_obj)

                        if len(areas_with_relations) > 0:
                            main_area = areas_with_relations[0]

                            # Update the name of the main area
                            main_area.areas = area_name.title()
                            main_area.save()

                            # Redirect relations from all other areas with relations to main_area
                            for area_obj in areas_with_relations[1:]:
                                with transaction.atomic():
                                    area_obj.familyheaddeatils_area.update(area=main_area)
                                    area_obj.delete()

                            # Delete all areas without relations
                            for area_obj in areas_without_relations:
                                area_obj.delete()

                            self.stdout.write(self.style.SUCCESS(f"Update successful for area: {main_area.areas}"))

                        else:
                            # No areas have relations, so keep the first one and delete the rest
                            main_area = similar_areas[0]

                            # Update the name of the main area
                            main_area.areas = area_name.title()
                            main_area.save()

                            for area_obj in similar_areas[1:]:
                                area_obj.delete()

                            self.stdout.write(self.style.SUCCESS(f"Update successful for area: {main_area.areas}"))

                    elif len(similar_areas) == 1:
                        main_area = similar_areas[0]

                        # Update the name of the main area
                        main_area.areas = area_name.title()
                        main_area.save()

                        self.stdout.write(self.style.SUCCESS(f"Update successful for area: {main_area.areas}"))
                    else:
                        # No similar areas, it might be an insert
                        area_name = area_name.title()
                        area.objects.create(areas=area_name, healthPost_id=healthpost_id, dispensary_id=dispensary_id)
                        self.stdout.write(self.style.SUCCESS(f"Insert successful for area: {area_name}"))

                elif action == 'Delete':
                    if similar_areas:
                        for area_obj in similar_areas:
                            if not area_obj.familyheaddeatils_area.exists():
                                area_obj.delete()
                            else:
                                self.error_excel_data.append([ward_name, healthpost_name, healthpost_id, dispensary_name, dispensary_id,
                                                              area_obj.areas, "Delete", "has relations in family head"])


    def process_areas(self, area_objs:list, sections:list) -> list:
        """
        Process the area objects with the list of areas in the excel and returns the areas,
        and the actions to be performed on that objects.

        Parameters:
        - sections (list): List of sections in the area mapping excel.
        - area_objs (area): Area objects that will be comapred against the area mapping.

        Retruns: None.
        """
        # A list containing the area and their respective actions.
        processed_areas = []

        for section in sections:
            for area_name in section["areas"]:
                area_objs_copy = area_objs.copy()
                area_name = area_name.strip().lower()

                area_dict = {"area_name":area_name,
                            "similar_areas":[],
                            "action":None}

                for area_obj in area_objs_copy:
                    db_area_name = area_obj.areas.lower()

                    if db_area_name == area_name:
                        area_dict["similar_areas"].append(area_obj)
                        area_dict["action"] = "NA"
                        area_objs.remove(area_obj)
                    elif db_area_name in area_name:
                        # print(area_name, db_area_name,"****")
                        area_dict["similar_areas"].append(area_obj)
                        area_dict["action"] = "Update"
                        area_objs.remove(area_obj)

                if not area_dict["similar_areas"]:
                    for area_obj in area_objs_copy:
                        db_area_name = area_obj.areas.lower()

                        if self.is_similar(db_area_name, area_name):
                            area_dict["similar_areas"].append(area_obj)
                            area_dict["action"] = "Update"
                            area_objs.remove(area_obj)

                if area_dict["similar_areas"]:
                    processed_areas.append(area_dict)
                else:
                    area_dict["action"] = "Insert"
                    processed_areas.append(area_dict)

        if area_objs:
            processed_areas.append({"area_name":"NA",
                                    "similar_areas":area_objs,
                                    "action":"Delete"})
        return processed_areas



    def is_similar(self, text1:str, text2:str) -> bool:
        """
        Calculates the Jaccard similarity between two texts and returns True if the similarity
        calculates above 0.8 else it will return False.

        Parameters:
        - text1: The first text.
        - text2: The second text.

        Returns: A boolean representing that is the two texts are similar or not.
        """
        set1 = set(text1.lower())
        set2 = set(text2.lower())

        intersection = set1 & set2
        union = set1 | set2

        jaccard_similarity = len(intersection)/len(union)

        return 1 if jaccard_similarity >= 0.8 else 0