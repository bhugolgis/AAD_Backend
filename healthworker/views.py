from rest_framework import generics
from rest_framework.response import Response
from database.models import *
from .serializer import *
from django.contrib.gis.geos import Point
from rest_framework import status
from rest_framework.parsers import MultiPartParser , JSONParser
from rest_framework.permissions import IsAuthenticated
import random
from .permissions import IsHealthworker , IsCHV_ASHA
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework import filters
from django.utils import timezone
from openpyxl import load_workbook
from django.contrib.auth.models import Group
from datetime import datetime
from django.db.models import Count, Q, Case, When, IntegerField, F, Value
from django.db.models.functions import Cast, Substr, Coalesce
from django.db.models.expressions import Func
import openpyxl
from django.http import HttpResponse
from ArogyaAplyaDari.utils import error_simplifier
from adminportal.utils import get_aggregated_data


class verifyMobileNumber(APIView):
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    def get(self , request ,mobileNo):
        """
        The function checks if a given mobile number already exists in the database and returns a
        response indicating whether it exists or not.

        :param request: The request object contains information about the current HTTP request, such as
        the headers, body, and method
        :param mobileNo: The mobileNo parameter is the mobile number that is being checked for existence
        in the familyHeadDetails and CustomUser models
        :return: a Response object with a status and message. If the mobile number already exists in
        either the familyHeadDetails or CustomUser models, it returns an error message with a status
        code of 400. Otherwise, it returns a success message with a status code of 200.
        """
        family_head = familyHeadDetails.objects.filter( mobileNo = mobileNo).exists()
        user = CustomUser.objects.filter(phoneNumber = mobileNo).exists()
        if family_head or user:
            return Response({'status' : 'error' ,
                            'message' : 'Mobile Number Already exist' } , status= status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'status' : 'success' ,
                            'message' : 'verify successfully' } , status= status.HTTP_200_OK)

class veirfyaadharCard(APIView):
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    def get(self , request ,aadharCard):
        """
        The function checks if a given Aadhar Card number already exists in the database and returns a
        response accordingly.

        :param request: The request object contains information about the current HTTP request, such as
        the headers, method, and body
        :param aadharCard: The parameter "aadharCard" is the Aadhar card number that is being passed to
        the function. It is used to check if a family member with the same Aadhar card number already
        exists in the database
        :return: The code is returning a response in the form of a dictionary. If the data exists, it
        returns a response with a status of 'error' and a message stating that the Aadhar Card already
        exists. If the data does not exist, it returns a response with a status of 'success' and a
        message stating that the verification was successful.
        """
        data = familyMembers.objects.filter( aadharCard = aadharCard).exists()
        if data:
            return Response({'status' : 'error' ,
                            'message' : 'Aadhar Card Already exist' } , status= status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'status' : 'success' ,
                            'message' : 'verify successfully' } , status= status.HTTP_200_OK)

class verifyabhaId(APIView):
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    def get(self , request ,abhaId):
        """
        The function checks if a given abhaId already exists in the familyMembers table and returns a
        response indicating whether it exists or not.

        """
        data = familyMembers.objects.filter( abhaId = abhaId).exists()
        if data:
            return Response({'status' : 'error' ,
                            'message' : 'ABHA-ID Already exist' } , status= status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'status' : 'success' ,
                            'message' : 'verify successfully' } , status= status.HTTP_200_OK)

class PostSurveyForm(generics.GenericAPIView):
    serializer_class = PostSurveyFormSerializer
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    parser_classes = [JSONParser]

    def post(self , request , *args , **kwargs):
        """
        The above function is a view function in Django that handles the POST request for saving data,
        including validating and saving the data, generating a family ID, and returning a response.

        """
        serializer = self.get_serializer(data = request.data)
        # print(serializer)
        if serializer.is_valid():
            try:
                lat = float(serializer.validated_data['latitude'] , None)
                long = float(serializer.validated_data['longitude'] , None)
                location = Point(long, lat, srid=4326)
            except:
                location = None

            user =  request.user
            ward =  (user.userSections.all()[0].healthPost.ward.wardName).replace(" ","")
            familyId = 'F-{ward}-{num}'.format(ward = ward, num = random.randint(0000000 , 9999999))
            serializer.save(location = location , familyId = familyId ,user = user )
            return Response ({'status' : 'success' ,
                            'message' : 'data saved successfully' } , status= status.HTTP_200_OK)
        else:
            print(serializer.errors)
            key, value =list(serializer.errors.items())[0]
            try:
                key , value = list(value[0].items())[0]
                error_message =  str(value[0])
            except Exception as e:
                try:
                     error_message = str(key) + " ," +str(value[0])
                except:
                    key , value = list(value[1].items())[0]
                    error_message = str(key) + " ," +str(value[0])

            return Response({'status': 'error',
                            'message' :error_message} , status = status.HTTP_400_BAD_REQUEST)


# The class `GetFamilyHeadList` is a generic ListAPIView that retrieves a list of family head details
# and allows filtering by mobile number, name, and family ID.
class GetFamilyHeadList(generics.ListAPIView):
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    serializer_class = GetFamilyHeadListSerialzier
    filter_backends = (filters.SearchFilter,)
    search_fields = ['mobileNo', 'name', 'familyId']


    def get_queryset(self):
        usersection =  self.request.user.userSections.all()
        section_ids = [section.id for section in usersection]
        user_group =self.request.user.groups.all()[0]

        if str(user_group) == 'CHV-ASHA':
            # Filter the queryset based on the currently logged-in user
            queryset = familyHeadDetails.objects.filter(user=self.request.user).order_by('-created_date')
        elif str(user_group) == "healthworker":
            queryset = familyHeadDetails.objects.filter(user__userSections__id__in=section_ids).distinct().order_by('-created_date')
        return queryset



class GetPartiallyInsertedRecord(generics.ListAPIView):
    permission_classes =(IsAuthenticated, IsHealthworker | IsCHV_ASHA)
    serializer_class = GetFamilyHeadListSerialzier
    filter_backends = (filters.SearchFilter,)
    search_fields = ['mobileNo', 'name', 'familyId']

    def get_queryset(self):
        usersection =  self.request.user.userSections.all()
        section_ids = [section.id for section in usersection]
        user_group =self.request.user.groups.all()[0]
        if str(user_group) == 'CHV-ASHA':
            # Filter the queryset based on the currently logged-in user
            queryset = familyHeadDetails.objects.filter(user=self.request.user, partialSubmit=True).order_by('-created_date')
        elif str(user_group) == "healthworker":
            queryset = familyHeadDetails.objects.filter(user__userSections__id__in=section_ids, partialSubmit=True).distinct().order_by('-created_date')
        return queryset

class PostFamilyDetails(generics.GenericAPIView):
    serializer_class = postFamilyMemberDetailSerializer
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    def post(self , request , *args , **kwargs):
        """
        This function saves data from a serializer and returns a success message if the data is valid,
        or an error message if the data is invalid.
        """
        serializer = self.get_serializer(data = request.data)
        if serializer.is_valid():
            # print(serializer.validated_data)
            get_last_memberId = familyMembers.objects.filter(familyHead_id = serializer.validated_data['familyHead']).last()
            if get_last_memberId:
                last_member_id = get_last_memberId.memberId
                last_member_id_parts = last_member_id.split("-")

                numeric_part = int(last_member_id_parts[-1])
                new_numeric_part = numeric_part + 1

                new_member_id = "-".join(last_member_id_parts[:-1]) + "-" + str(new_numeric_part).zfill(2)
            else:
                new_member_id = "F-F/S-7145623-01"
            serializer.save(familySurveyor = request.user  , memberId = new_member_id)
            return Response ({'status' : 'success' ,
                            'message' : 'data saved successfully' } , status= status.HTTP_200_OK)
        else:
            error_message = error_simplifier(serializer.errors)
            return Response({"status" : "error" ,
                             "message" : error_message} , status= status.HTTP_400_BAD_REQUEST)

# The class `GetFamilyMembersDetails` is a generic ListAPIView that retrieves details of family
# members and allows filtering based on family ID, mobile number, and name.
class GetFamilyMembersDetails(generics.ListAPIView):
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    serializer_class = GetFamilyMemberDetailSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ['familyHead__familyId' , 'familyHead__mobileNo' , 'familyHead__name' , 'memberId'  ]

    def get_queryset(self):
        usersection =  self.request.user.userSections.all()[0]
        user_group =self.request.user.groups.all()[0]
        if str(user_group) == 'CHV-ASHA':
            # Filter the queryset based on the currently logged-in user
            queryset = familyMembers.objects.filter( area__healthPost__ward = usersection.healthPost.ward , familySurveyor=self.request.user).order_by('-created_date')
        elif str(user_group) == "healthworker":
            queryset = familyMembers.objects.filter(area__healthPost__ward = usersection.healthPost.ward ,  area__healthPost__healthPost_name = usersection ).order_by('-created_date')
        return queryset


class UpdateFamilyDetails(generics.GenericAPIView):

    serializer_class = UpdateFamilyMemberDetailSerializer
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    # parser_classes = [MultiPartParser]

    def patch(self , request , id , *args , **kwargs):
        """
        The function `patch` updates an instance of the `familyMembers` model with the provided `id`
        using the data from the `request`, and returns a success message if the update is successful, or
        an error message if there are any validation errors.
        """
        try:
            instance = familyMembers.objects.get(id =id)
        except:
            return Response({'status': 'error',
                             'message': 'caseNumber ID is not found'}, status=400)

        serializer = self.get_serializer( instance , data = request.data , partial = True)
        if serializer.is_valid():
            serializer.save(familySurveyor = request.user )
            return Response ({'status' : 'success' ,
                            'message' : 'data saved successfully' } , status= status.HTTP_200_OK)
        else:
            key, value =list(serializer.errors.items())[0]
            error_message = key+" , "+ value[0]
            return Response({"status" : "error" ,
                             "message" : error_message} , status= status.HTTP_400_BAD_REQUEST)


class GetSurveyorCountDashboard(generics.GenericAPIView): # Modified
    permission_classes = (IsAuthenticated, IsHealthworker | IsCHV_ASHA)
    queryset = familyMembers.objects.all()
    FamilySurvey_count = familyHeadDetails.objects.all()
    serializer_class = GetFamilyMemberDetailSerializer

    def get(self, request,  *args, **kwargs):
        """
        This function returns various counts related to citizen and family surveys for a specific user.
        """
        usersection =  self.request.user.userSections.all()
        section_ids = [section.id for section in usersection]
        user_group =self.request.user.groups.all()[0]
        today = timezone.now().date()

        if str(user_group) == 'CHV-ASHA':
            # Distinct and common queries of survey data
            survey_queryset = self.get_queryset().filter(familySurveyor=request.user).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user=request.user)

            # Healthppost related survey data
            aggregated_data = get_aggregated_data(survey_queryset, familySurvey_queryset)

        elif str(user_group) == 'healthworker':
            # Distinct and common queries of survey data
            survey_queryset = self.get_queryset().filter(familySurveyor__userSections__id__in=section_ids).distinct()
            familySurvey_queryset = self.FamilySurvey_count.filter(user__userSections__id__in=section_ids)

            # Healthppost related survey data
            aggregated_data = get_aggregated_data(survey_queryset, familySurvey_queryset)

        return Response(aggregated_data, status=status.HTTP_200_OK)


class GetCitizenList(generics.ListAPIView):
    serializer_class = GetCitizenListSerializer
    model = serializer_class.Meta.model
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    filter_backends = (filters.SearchFilter,)
    search_fields = ['familyHead__familyId' , 'familyHead__mobileNo' , 'familyHead__name' , 'memberId' , 'name' , 'mobileNo' ]

    def get_queryset(self ):
        """
        The function returns a queryset of all objects ordered by their created date in descending order.
        """
        queryset = self.model.objects.all()
        return queryset.order_by('-created_date')

    def get(self , request , choice):
        """
        The function `get` retrieves data based on the user's choice of either "Today", "All", or
        "Partially".
        :param request: The `request` parameter is an object that represents the HTTP request made by the
        client. It contains information such as the user making the request, the method used (GET, POST,
        etc.), and any data or parameters sent with the request
        :param choice: The `choice` parameter is a string that determines the type of data to be fetched.
        It can have three possible values: "Today", "All", or "Partially"
        :return: a Response object with a message, status, and data. The message indicates that the data
        was fetched successfully, the status indicates the success status, and the data contains the
        serialized queryset. The status code of the response is HTTP 200 OK.
        """
        usersection =  self.request.user.userSections.all()[0]
        user_group =self.request.user.groups.all()[0]
        if choice == 'Today':
            today = timezone.now().date()
            if str(user_group) == 'CHV-ASHA':
                queryset = self.model.objects.filter(area__healthPost__ward = usersection.healthPost.ward , familySurveyor = request.user , created_date__day = today.day).order_by('-created_date')
            elif str(user_group) == "healthworker":
                 queryset = self.model.objects.filter(area__healthPost__ward = usersection.healthPost.ward ,  area__healthPost__healthPost_name = usersection , created_date__day = today.day ).order_by('-created_date')
            serializer = self.get_serializer(queryset , many = True ).data
            return Response(serializer , status= status.HTTP_200_OK)

        elif choice == 'All':
            if str(user_group) == 'CHV-ASHA':
                total_list = self.get_queryset().filter(area__healthPost__ward = usersection.healthPost.ward ,familySurveyor = request.user).order_by('-created_date')
            elif str(user_group) == "healthworker":
                total_list = self.get_queryset().filter(area__healthPost__ward = usersection.healthPost.ward ,  area__healthPost__healthPost_name = usersection ).order_by('-created_date')

            serializer = self.get_serializer(total_list , many = True ).data
            return Response(serializer  , status= status.HTTP_200_OK)

        elif choice == 'Partially':
            if str(user_group) == 'CHV-ASHA':
                queryset = familyHeadDetails.objects.filter(area__healthPost__ward = usersection.healthPost.ward ,partialSubmit = True , user = request.user).order_by('-created_date')
            elif str(user_group) == "healthworker":
                queryset = familyHeadDetails.objects.filter(area__healthPost__ward = usersection.healthPost.ward ,partialSubmit = True , area__healthPost__healthPost_name = usersection ).order_by('-created_date')
            serializer = GetFamilyHeadListSerialzier(queryset , many = True ).data
            return Response({ serializer } , status= status.HTTP_200_OK)


class getReferelOptionList(generics.ListAPIView):
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    serializer_class = getReferelOptionListSerialzier
    model = serializer_class.Meta.model
    queryset = model.objects.all()

class getvulnerableOptionList(generics.ListAPIView):
    serializer_class = getvulnerableOptionListSerialzier
    model = serializer_class.Meta.model
    queryset = model.objects.all()

class GetFamilyList(generics.ListAPIView):
    serializer_class = GetFamilyHeadListSerialzier
    model = serializer_class.Meta.model
    permission_classes = (IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    filter_backends = (filters.SearchFilter,)
    search_fields = ['familyHead__familyId' , 'familyHead__mobileNo' , 'familyHead__name' , 'memberId' , 'name' , 'mobileNo' ]
    paginate_by = 100

    def get_queryset(self ):
        queryset = self.model.objects.all()
        return queryset.order_by('-created_date')

    def get(self , request , choice):
        """
        The function `get` retrieves data based on the user's choice of either "Today" or "All" and
        returns a response with the fetched data.

        :param request: The `request` parameter is an object that represents the HTTP request made by
        the client. It contains information such as the request method (GET, POST, etc.), headers, user
        authentication, and query parameters
        :param choice: The `choice` parameter is used to determine the type of data to fetch. It can
        have two possible values: "Today" or "All"
        :return: The code is returning a response object with a message, status, and data. The message
        indicates that the data was fetched successfully, the status indicates success, and the data
        contains the serialized queryset. The status code of the response is 200 (OK).
        """

        if choice == 'Today':
            today = timezone.now().date()
            usersection =  self.request.user.userSections.all()[0]
            user_group =self.request.user.groups.all()[0]
            if str(user_group) == 'CHV-ASHA':
                queryset = self.model.objects.filter(area__healthPost__ward = usersection.healthPost.ward ,user = request.user , created_date__day = today.day).order_by('-created_date')
            elif str(user_group) == "healthworker":
                queryset = self.model.objects.filter(area__healthPost__ward = usersection.healthPost.ward ,  area__healthPost__healthPost_name = usersection ).order_by('-created_date')

            serializer = self.get_serializer(queryset , many = True ).data
            return Response( serializer  , status= status.HTTP_200_OK)

        elif choice == 'All':
            if str(user_group) == 'CHV-ASHA':
                total_list = self.get_queryset().filter(area__healthPost__ward = usersection.healthPost.ward ,user = request.user).order_by('-created_date')
            elif str(user_group) == "healthworker":
                total_list = self.get_queryset().filter(area__healthPost__ward = usersection.healthPost.ward ,  area__healthPost__healthPost_name = usersection ).order_by('-created_date')

            serializer = self.get_serializer(total_list , many = True ).data
            return Response( serializer , status= status.HTTP_200_OK)


class GetBloodCollectionDetail(generics.ListAPIView):
    queryset = familyMembers.objects.all().order_by('-created_date')
    serializer_class = GetCitizenListSerializer
    permission_classes =(IsAuthenticated , IsHealthworker | IsCHV_ASHA)
    filter_backends = (filters.SearchFilter,)
    search_fields = ['bloodCollectionLocation']

class UserInsertxlsx(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = DumpExcelSerializer

    def post(self, request, format=None):
        """
        The above function is a Django view that handles the uploading of an Excel file, reads the data
        from the file, and creates user objects in the database based on the data.

        :param request: The `request` parameter is an object that contains information about the current
        HTTP request. It includes details such as the request method (GET, POST, etc.), headers, body,
        and any uploaded files
        :param format: The `format` parameter is used to specify the desired format of the response. In
        this case, it is set to `None`, which means that the format will be determined automatically
        based on the request
        :return: The code is returning a response object with a JSON payload. The payload contains a
        "message" field indicating the status of the operation ("File Uploaded Successfully and users
        created !!" or "Something Went wrong please check you File"), an "error" field containing the
        error message if an exception occurred, and a "status" field indicating the status of the
        operation ("success" or "error"). The status
        """
        try:
            if 'excel_file' not in request.FILES:
                return Response({"status": "error", "message": "No file uploaded."}, status=400)
            excel_file = request.FILES["excel_file"]

            if excel_file.size == 0 or excel_file.name.endswith(".xlsx") != True:
                return Response({"status": "error",
                                "message": "only .xlsx file is supported."},
                                status=400)

            workbook = load_workbook(filename=excel_file)
            sheet_name = workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                already_exist = CustomUser.objects.filter(username =row[1] , phoneNumber=row[3]  ).exists()
                if already_exist:
                    continue
                user = CustomUser.objects.create_user(name = row[0] , username=row[1],
                                                      password=row[2], phoneNumber=row[3],section_id=row[4])
                group = Group.objects.get(name = 'CHV-ASHA')
                user.groups.add(group)

            return Response({'message' : 'File Uploaded Successfully and users created !!' ,
                            'status' :"success"} , status= status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'message' : 'Something Went wrong please check you File',
                             'error' : str(e),
                            'status' :"error"} , status= status.HTTP_400_BAD_REQUEST)


class DownloadANM_CHV_UserList(generics.GenericAPIView):
    permission_classes = [IsAuthenticated , IsCHV_ASHA | IsHealthworker ]

    def add_headers(self, sheet, *args):
        for header in range(len(args)):
            if isinstance(args[header],dict):
                start_column = 1
                for title,size in args[header].items():
                    end_column = start_column + (size-1)
                    sheet.merge_cells(start_row=header+1,start_column=start_column,
                                      end_row=header+1, end_column=end_column)
                    sheet.cell(row=header+1, column=start_column, value=title)
                    start_column = end_column + 1
            else:
                sheet.append(args[header])
        return sheet

    def unpack_list(self, data):
        val = ""
        if len(data) == 1:
            val = data[0]
        elif len(data) > 1:
            for i in data:
                val = val + i + ", "
        return val

    def unpack_survey_data(self, survey_data):
        collected_data = []

        for data in survey_data.values():
            for answers in data:
                answer = answers.get("answer",None)
                collected_data.append(self.unpack_list(answer))

        return collected_data

    def get(self, request, *args, **kwargs):

        try:
            usersection =  request.user.userSections.all()[0]
            user_group =request.user.groups.all()[0]
            if str(user_group) == "healthworker":
                related_user = familyMembers.objects.filter(area__healthPost__ward = usersection.healthPost.ward ,  area__healthPost__healthPost_name = usersection).order_by('-created_date')
            elif str(user_group) == 'CHV-ASHA':
                related_user = familyMembers.objects.filter( area__healthPost__ward = usersection.healthPost.ward , familySurveyor=self.request.user).order_by('-created_date')
        except familyMembers.DoesNotExist:
            return Response({
                "message":"No data found",
                "status":"error"
            } , status=400)

        today = datetime.today().strftime('%d-%m-%Y')
        # healthpost_name = healthpost.healthPostName

        if not related_user:
            return Response({
                "message":"No data found",
                "status":"error"
            } , status=400)

        familyMember = related_user.last()
        questionnaire = familyMember.Questionnaire
        parts_dict = {}
        questions_list = []
        for part,questions in questionnaire.items():
            parts_dict[part] = len(questions)
            for question in questions:
                questions_list.append(question.get("question",None))

        column_list = ['Name', 'Gender', 'Age', 'Mobile No', "Address" ,  'Aadhar Card', 'ABHA ID',
                       'Blood Collection Location', 'Family Head', 'ANM/Coordinator', 'ANM/Coordinator Mobile Number' , 'Survey Date',
                       'BMI', 'Blood Pressure', 'Height', 'Pulse', 'Weight', 'Test Assigned',
                       'Report', 'Area', 'General Status', 'ASHA/CHV', 'ASHA/CHV Mobile Number' , 'Vulnerable',
                       'Vulnerable Reason', 'Relationship', 'Random Blood Sugar']

        header1 = {'Citizen Details':len(column_list),
                   'Survey':len(questions_list)}
        header2 = {'':len(column_list),**parts_dict}
        header3 = column_list + questions_list

        data_list = []
        for family_member in  related_user:
            citizen_details = [family_member.name, family_member.gender, family_member.age, family_member.mobileNo, family_member.familyHead.address,
                               family_member.aadharCard, family_member.abhaId,
                               family_member.bloodCollectionLocation, family_member.familyHead.name,
                               family_member.familySurveyor.name, family_member.familySurveyor.phoneNumber , family_member.created_date.strftime('%d/%m/%Y'), family_member.BMI,
                               family_member.bloodPressure, family_member.height, family_member.pulse,
                               family_member.weight, family_member.bool_transform("isLabTestAdded"),
                               family_member.bool_transform("isLabTestReportGenerated"),
                               family_member.area.areas if family_member.area else None,
                               family_member.generalStatus, family_member.ASHA_CHV.name if family_member.ASHA_CHV else None , family_member.ASHA_CHV.phoneNumber if family_member.ASHA_CHV else None ,
                               family_member.bool_transform("vulnerable"), family_member.vulnerable_reason, family_member.relationship,
                               family_member.randomBloodSugar]
            survey_data = self.unpack_survey_data(family_member.Questionnaire)
            aggregated_data = citizen_details + survey_data
            data_list.append(aggregated_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        self.add_headers(ws, header1, header2, header3)
        for row in data_list:
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format("data_"+today)
        wb.save(response)
        return response




